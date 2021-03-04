import xlrd
import xlwt
import openpyxl
from multiprocessing.dummy import Pool as ThreadPool

PATH = '/mnt/c/Users/hikki/OneDrive/210107/word/'


# get excel booksheet form path
def getExcelToSheet(path, sheetIndex=0):
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    worksheet = workbook.sheet_by_index(sheetIndex)
    return worksheet


# get XLSX excel booksheet form path
def getExcelToSheetXLSX(path, sheetname):
    workbook = openpyxl.load_workbook(path, read_only=True)  # 打开工作簿
    worksheet = workbook[sheetname]
    return worksheet


# get sheet row
def getSheetRow(sheet):
    return sheet.nrows


# get XLSX sheet row
def getSheetRowXLSX(sheet):
    return sheet.max_row


# get sheet column
def getSheetCol(sheet):
    return sheet.ncols


# get XLSX sheet column
def getSheetColXLSX(sheet):
    return sheet.max_col


# get the first line
def getFirstLine(path, sheetIndex=0):
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_by_index(sheetIndex)
    return worksheet.row_values(0)


# get cell value from sheet
def getValue(sheet, row, col):
    return sheet.cell(row, col).value


# save xls from array
def saveExcel(array, filename='res.xls'):
    wbk = xlwt.Workbook(encoding="utf-8")
    datasheet = wbk.add_sheet("sheet1")
    for row in range(len(array)):
        for col in range(len(array[row])):
            datasheet.write(row, col, array[row][col])
    wbk.save(filename)


# save xls from dict
def saveExcelFromDict(dic, filename='res.xls'):
    wbk = xlwt.Workbook(encoding="utf-8")
    datasheet = wbk.add_sheet("sheet1")
    row = 0
    for code in dic:
        for year in dic[code]:
            col = 0
            datasheet.write(row, col, code)
            col += 1
            datasheet.write(row, col, year)
            col += 1
            datasheet.write(row, col, dic[code][year])
            row += 1
    wbk.save(filename)

    wbk.save(filename)


# get an array from row
def getRow(row):
    row_list = []
    for cell in row:
        row_list.append(cell.value)
    return row_list


# get a dict from XLSX
def getDicFromSheetXLSX(sheet):
    pool = ThreadPool()
    # total_list = []
    # for row in sheet.rows:
    #     row_list=getRow(row)
    #     total_list.append(row_list)
    total_list = pool.map(getRow, sheet.rows)
    pool.close()
    pool.join()
    total_dic = {}
    for name in range(1, getSheetRowXLSX(sheet)):
        total_dic.setdefault(name, {})
        for subname in total_list[0]:
            total_dic[name].setdefault(subname, 0)
            total_dic[name][subname] = total_list[name][total_list[0].index(subname)]
    return total_dic


if __name__ == '__main__':
    firstLine = getFirstLine(PATH + 'target.xls')
    codeSheet = getExcelToSheet(PATH + 'code.xls')
    gdpSheet = getExcelToSheet(PATH + 'gdp.xls')
    invSheet = getExcelToSheet(PATH + 'invest_oecd_sum.xls')
    tradeSheet = getExcelToSheetXLSX(PATH + 'WB-tradecosts-dataset.xlsx', 'GTT')
    dic = getDicFromSheetXLSX(tradeSheet)
    res = [firstLine]
    cnt = 0
    for pointer1 in range(1, getSheetRow(codeSheet)):
        code1 = getValue(codeSheet, pointer1, 0)
        country1 = getValue(codeSheet, pointer1, 1)
        region1 = getValue(codeSheet, pointer1, 2)
        for pointer2 in range(pointer1, getSheetRow(codeSheet)):
            code2 = getValue(codeSheet, pointer1, 0)
            country2 = getValue(codeSheet, pointer1, 1)
            region2 = getValue(codeSheet, pointer1, 2)
            for year in range(2000, 2019):
                gdp1 = 0
                gdp2 = 0
                inv1 = 0
                inv2 = 0
                tij = 0
                tar = 0
                for row in range(4, getSheetRow(gdpSheet)):
                    if code1 == getValue(gdpSheet, row, 1):
                        gdp1 = getValue(gdpSheet, row, year - 1960 + 4)
                        continue
                    if code2 == getValue(gdpSheet, row, 1):
                        gdp2 = getValue(gdpSheet, row, year - 1960 + 4)
                        continue
                    if gdp1 != 0 and gdp2 != 0:
                        break
                for row in range(1, getSheetRow(invSheet)):
                    if code1 == getValue(invSheet, row, 0) \
                            and year == getValue(invSheet, row, 1):
                        inv1 = getValue(invSheet, row, 2)
                        continue
                    if code2 == getValue(invSheet, row, 0) \
                            and year == getValue(invSheet, row, 1):
                        inv2 = getValue(invSheet, row, 2)
                        continue
                    if inv1 != 0 and inv2 != 0:
                        break
                for index in dic:
                    if code1 == dic[index]['reporter']:
                        if code2 == dic[index]['partner']:
                            if year == dic[index]['year']:
                                tij = dic[index]['tij']
                                tar = dic[index]['geometric_avg_tariff']
                                break
                            continue
                        continue
                    continue
                newRecord = [country1, code1, region1, country2, code2, region2, year,
                             gdp1, gdp2, inv1, inv2, tij, tar]
                for col in range(len(newRecord)):
                    if newRecord[col] == 0:
                        newRecord[col] = ''
                res.append(newRecord)
                print(cnt)
                cnt += 1
    saveExcel(res)
